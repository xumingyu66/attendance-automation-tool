"""Generate a blank attendance template based on employee data from the database.

Matches the structure of 文件考勤.xlsx with repeating headers every 10 employees.
"""

from __future__ import annotations

import calendar
import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from wrappers.db_handler import DatabaseHandler


class TemplateGenerator:
    """Generates a blank attendance template xlsx from DB employee data."""

    DAY_COL_START = 3
    DAY_COL_END = 33
    TOTAL_COL = 34
    TITLE_ROW = 1
    HEADER_ROW = 2
    WEEKDAY_ROW = 3

    WEEKDAY_NAMES = ["一", "二", "三", "四", "五", "六", "日"]
    REPEAT_INTERVAL = 10  # insert repeating header every N employees

    def __init__(self, db_dir: str) -> None:
        self._db = DatabaseHandler(db_dir)
        self._header_rows: list[int] = []   # rows that are repeating headers
        self._weekday_rows: list[int] = []  # rows that are repeating weekday labels
        self._last_row = 0

    def generate(
        self,
        output_path: str,
        year: int,
        month: int,
    ) -> int:
        """Generate a blank attendance template."""
        employees = self._db.get_all_employees()
        if not employees:
            raise ValueError("数据库中没有员工信息，请先导入员工数据")

        wb = Workbook()
        ws = wb.active
        ws.title = "工作时间统计"

        days_in_month = calendar.monthrange(year, month)[1]

        self._write_main_header(ws, year, month, days_in_month)
        self._write_body(ws, employees, year, month, days_in_month)
        self._apply_styles(ws, days_in_month)

        saved_path = self._safe_save(wb, output_path)
        self._last_saved_path = saved_path
        return len(employees)

    @property
    def last_saved_path(self) -> str:
        return getattr(self, "_last_saved_path", "")

    def _safe_save(self, wb: Workbook, path: str) -> str:
        from pathlib import Path

        base = Path(path)
        candidates = [base] + [
            base.with_stem(f"{base.stem}{i:02d}") for i in range(1, 100)
        ]
        for p in candidates:
            try:
                wb.save(str(p))
                return str(p)
            except PermissionError:
                continue
        raise PermissionError(f"无法保存文件，所有候选路径均被占用: {path}")

    # =========================================================================
    # Main header (top of sheet)
    # =========================================================================

    def _write_main_header(self, ws, year: int, month: int, days_in_month: int) -> None:
        max_col = self.TOTAL_COL

        # R1: title
        title = f"{year}年{month}月考勤统计表"
        ws.cell(row=self.TITLE_ROW, column=1, value=title)
        ws.merge_cells(
            start_row=self.TITLE_ROW, start_column=1,
            end_row=self.TITLE_ROW, end_column=max_col,
        )

        self._write_header_row(ws, self.HEADER_ROW, self.WEEKDAY_ROW, year, month, days_in_month)
        self._header_rows.append(self.HEADER_ROW)
        self._weekday_rows.append(self.WEEKDAY_ROW)

    # =========================================================================
    # Body: employees with repeating headers every N employees
    # =========================================================================

    def _write_body(self, ws, employees, year: int, month: int, days_in_month: int) -> None:
        row = 4  # first data row after the main header

        for i, emp in enumerate(employees):
            # Insert repeating header before this employee if at interval boundary
            if i > 0 and i % self.REPEAT_INTERVAL == 0:
                header_r = row
                weekday_r = row + 1
                self._write_header_row(ws, header_r, weekday_r, year, month, days_in_month)
                self._header_rows.append(header_r)
                self._weekday_rows.append(weekday_r)
                row += 2

            ws.cell(row=row, column=1, value=str(emp.employee_number))
            ws.cell(row=row, column=2, value=emp.name)
            ws.cell(row=row, column=self.TOTAL_COL, value=0)
            row += 1

        self._last_row = row - 1

    # =========================================================================
    # Header row helper (used for both main header and repeating headers)
    # =========================================================================

    def _write_header_row(self, ws, header_r: int, weekday_r: int, year: int, month: int, days_in_month: int) -> None:
        """Write a header row pair at the given rows."""
        # Header: 工号 | 姓名 | 1-31 | 月总工时
        ws.cell(row=header_r, column=1, value="工号")
        ws.cell(row=header_r, column=2, value="姓名")
        for d in range(1, days_in_month + 1):
            col = self.DAY_COL_START + d - 1
            ws.cell(row=header_r, column=col, value=d)
        for d in range(days_in_month + 1, 32):
            col = self.DAY_COL_START + d - 1
            ws.cell(row=header_r, column=col, value=d)
        ws.cell(row=header_r, column=self.TOTAL_COL, value="月总工时")

        # Merge header+weekday for fixed columns
        ws.merge_cells(start_row=header_r, start_column=1, end_row=weekday_r, end_column=1)
        ws.merge_cells(start_row=header_r, start_column=2, end_row=weekday_r, end_column=2)
        ws.merge_cells(start_row=header_r, start_column=self.TOTAL_COL, end_row=weekday_r, end_column=self.TOTAL_COL)

        # Weekday labels
        for d in range(1, days_in_month + 1):
            col = self.DAY_COL_START + d - 1
            weekday = calendar.weekday(year, month, d)
            ws.cell(row=weekday_r, column=col, value=self.WEEKDAY_NAMES[weekday])

    # =========================================================================
    # Styling
    # =========================================================================

    def _apply_styles(self, ws, days_in_month: int) -> None:
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")
        title_font = Font(name="宋体", size=14, bold=True)
        header_font = Font(name="宋体", size=10, bold=True)
        body_font = Font(name="宋体", size=10)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        max_col = self.TOTAL_COL
        last_row = self._last_row

        # Borders + alignment for all cells
        for r in range(1, last_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
                cell.alignment = center_align

        # Title style
        title_cell = ws.cell(row=self.TITLE_ROW, column=1)
        title_cell.font = title_font
        title_cell.fill = header_fill

        # All header + weekday rows
        all_header_rows = set(self._header_rows) | set(self._weekday_rows)
        for r in all_header_rows:
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = header_font
                cell.fill = header_fill

        # Body font (all rows that aren't title or header/weekday)
        for r in range(1, last_row + 1):
            if r == self.TITLE_ROW or r in all_header_rows:
                continue
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).font = body_font

        # Column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 10
        for d in range(1, 32):
            col_letter = get_column_letter(self.DAY_COL_START + d - 1)
            ws.column_dimensions[col_letter].width = 4.5
        ws.column_dimensions[get_column_letter(self.TOTAL_COL)].width = 10

        # Row heights
        ws.row_dimensions[self.TITLE_ROW].height = 30
        for r in all_header_rows:
            ws.row_dimensions[r].height = 20
