"""Typed wrapper for openpyxl — reads and writes 文件考勤.xlsx."""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class AttendanceRow:
    row_number: int
    name: str
    daily_hours: dict[int, int | None]  # day_of_month -> hours (2 or None)


class XlsxHandler:
    """Reads and writes the attendance summary workbook (文件考勤.xlsx).

    Layout:
      - Row 1: metadata
      - Row 2: headers (序号 | 姓名 | 1 | 2 | ... | 31 | 月总工时)
      - Row 3: weekdays (四 | 五 | ...)
      - Row 4+: employee data rows

    The weekday row (R3) and month metadata (R1C1) are updated for the target month.
    """

    DAY_COL_START = 3   # Column C = day 1
    DAY_COL_END = 33    # Column AG = day 31 (max)
    TOTAL_COL = 34      # Column AH = 月总工时
    HEADER_ROW = 2
    WEEKDAY_ROW = 3
    DATA_START_ROW = 4

    WEEKDAY_NAMES = ["一", "二", "三", "四", "五", "六", "日"]

    def __init__(self, filepath: str) -> None:
        self._path = filepath
        self._workbook = openpyxl.load_workbook(filepath)
        self._sheet = self._workbook[self._workbook.sheetnames[0]]

    @property
    def max_data_row(self) -> int:
        return self._sheet.max_row

    def get_employees(self) -> list[AttendanceRow]:
        """Read all employee rows from the sheet."""
        rows: list[AttendanceRow] = []
        for r in range(self.DATA_START_ROW, self._sheet.max_row + 1):
            name = self._cell_str(r, 2)
            if not name:
                continue
            daily: dict[int, int | None] = {}
            for d in range(1, 32):
                val = self._cell_int(r, self.DAY_COL_START + d - 1)
                daily[d] = val if val else None
            rows.append(AttendanceRow(row_number=r, name=name, daily_hours=daily))
        return rows

    def update_weekdays(self, year: int, month: int) -> None:
        """Update the weekday row to match the given year/month."""
        import calendar

        days_in_month = calendar.monthrange(year, month)[1]

        for d in range(1, days_in_month + 1):
            weekday = calendar.weekday(year, month, d)  # 0=Mon ... 6=Sun
            day_col = self.DAY_COL_START + d - 1
            self._sheet.cell(row=self.WEEKDAY_ROW, column=day_col).value = self.WEEKDAY_NAMES[weekday]

        # Clear any extra days (if month has < 31 days)
        for d in range(days_in_month + 1, 32):
            day_col = self.DAY_COL_START + d - 1
            self._sheet.cell(row=self.WEEKDAY_ROW, column=day_col).value = None

        # Update month metadata in R1C1 (Excel date serial for first day)
        import datetime
        first_day = datetime.date(year, month, 1)
        epoch = datetime.date(1899, 12, 30)
        date_serial = (first_day - epoch).days
        self._sheet.cell(row=1, column=1, value=date_serial)

    def fill_hours(self, employee_name: str, daily_hours: dict[int, int | None]) -> None:
        """Fill attendance hours for a specific employee row.

        Args:
            employee_name: Name of the employee to match in column B.
            daily_hours: Dict of day_of_month -> hours (2) or None.
        """
        row_num = self._find_employee_row(employee_name)
        if row_num is None:
            return

        total = 0
        for d in range(1, 32):
            col = self.DAY_COL_START + d - 1
            cell = self._sheet.cell(row=row_num, column=col)
            if d in daily_hours and daily_hours[d] is not None:
                cell.value = daily_hours[d]
                total += daily_hours[d]
            else:
                cell.value = None
                # Also keep existing data if no data from attendance report

        self._sheet.cell(row=row_num, column=self.TOTAL_COL, value=total if total > 0 else 0)

    def _find_employee_row(self, name: str) -> int | None:
        """Find the row number for an employee by name."""
        for r in range(self.DATA_START_ROW, self._sheet.max_row + 1):
            cell_name = self._cell_str(r, 2)
            if cell_name == name:
                return r
        return None

    def save(self, output_path: str | None = None) -> None:
        path = output_path or self._path
        self._workbook.save(path)

    def _cell_str(self, row: int, col: int) -> str:
        val = self._sheet.cell(row=row, column=col).value
        if val is None:
            return ""
        return str(val).strip()

    def _cell_int(self, row: int, col: int) -> int | None:
        val = self._sheet.cell(row=row, column=col).value
        if val is None:
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None
